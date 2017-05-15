"""
SYNOPSIS

    make_primavera.py

DESCRIPTION

    Generate the additional PRIMAVERA mip tables for CMOR version 3. The tables
    are based on the data request, which has been saved as an Excel spreadsheet
    and is included in the Github repository:
    https://github.com/jonseddon/cmip6-cmor-tables

REQUIREMENTS

    Requires the openpyxl Python library. Tested with openpyxl version 2.4.1.
"""
from collections import OrderedDict
import json
import os

from openpyxl import load_workbook


EXCEL_FILE = 'PRIMAVERA_MS21_DRQ_0-beta-37.5.xlsx'

HEADER_COMMON = {
    'data_specs_version': '01.beta.45',
    'cmor_version': '3.2',
    'table_date': '19 December 2016',
    'missing_value': '1e20',
    'product': 'output',
    'generic_levels': '',
    'mip_era': 'CMIP6',
    'Conventions': 'CF-1.6 CMIP-6.0'
}


def generate_header(table_name):
    """
    Populate the `Header` section of the output dictionary with the
    variables in the PRIMAVERA mip table passed in to `req_sheet`.

    :param str table_name: The name of table being generated
    """
    # create a dict to hold the header
    header = {}

    # copy the standard header items into the header
    header.update(HEADER_COMMON)

    # set the table name
    header['table_id'] = 'Table {}'.format(table_name)

    # set the frequency
    # 1 hour should be 0.041666 but existing files use 0.017361 and so have
    # stuck with that
    frequencies = {'mon': '30.00000', 'day': '1.00000', '6hr': '0.250000',
                   '3hr': '0.125000', '1hr': '0.017361'}
    for freq in frequencies:
        if freq in table_name.lower():
            header['frequency'] = freq
            break

    if 'frequency' not in header:
        msg = ('Unable to determine frequency for table name: {}'.
               format(table_name))
        raise ValueError(msg)

    # create a blank realm, which will be populated later
    header['realm'] = ''

    # set the approx_interval
    header['approx_interval'] = frequencies[header['frequency']]

    ordered_keys = ['data_specs_version', 'table_id', 'realm', 'frequency',
                    'cmor_version', 'table_date', 'missing_value', 'product',
                    'approx_interval', 'generic_levels', 'mip_era',
                    'Conventions']
    key_order = {key: index for index, key in enumerate(ordered_keys)}

    return OrderedDict(sorted(header.items(),
                              key=lambda i: key_order.get(i[0])))


def generate_variable_entry(req_sheet, output):
    """
    Populate the `variable_entry` section of the `output` dictionary with the
    variables in the PRIMAVERA mip table passed in to `req_sheet`.

    :param openpyxl.worksheet.worksheet.Worksheet req_sheet: A worksheet from
        the Data Request spreadsheet
    :param dict output: The `variable_entry` dictionary from the output
        dictionary that will be converted to a JSON file
    """
    ordered_keys = ['modeling_realm', 'standard_name', 'units', 'cell_methods',
                    'cell_measures', 'long_name', 'comment', 'dimensions',
                    'out_name', 'type', 'positive', 'valid_min', 'valid_max',
                    'ok_min_mean_abs', 'ok_max_mean_abs']
    key_order = {key: index for index, key in enumerate(ordered_keys)}

    for row in req_sheet.iter_rows(min_row=2):
        var_name = _get_cell(row, 'var_name')
        if not var_name:
            # blank row so skip to next
            continue

        # create a blank dict for this variable
        var_dict = {}

        # get the CMOR name and if not defined use var_name
        cmor_name = _get_cell(row, 'cmor_name')
        if not cmor_name:
            cmor_name = var_name

        # these are the standard components that can be obtained directly from
        # the data request
        direct_components = ['modeling_realm', 'standard_name', 'units',
                             'cell_methods', 'cell_measures', 'long_name',
                             'comment', 'dimensions', 'type', 'positive']

        # add these standard components
        for cmpt in direct_components:
            cmpt_value = str(_get_cell(row, cmpt))
            if not cmpt_value:
                cmpt_value = ''
            var_dict[cmpt] = cmpt_value

        # fix the positive attribute
        if var_dict['positive'] == 'None':
            var_dict['positive'] = ''

        # add the remaining additional components
        var_dict['out_name'] = cmor_name
        var_dict['valid_min'] = ''
        var_dict['valid_max'] = ''
        var_dict['ok_min_mean_abs'] = ''
        var_dict['ok_max_mean_abs'] = ''

        # apply an order to this variable's dictionary and add it to the output
        output[cmor_name] = OrderedDict(
            sorted(var_dict.items(), key=lambda i: key_order.get(i[0])))


def main():
    """
    Run the software
    """
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.abspath(os.path.join(current_dir, '..',
                                              'etc', EXCEL_FILE))
    data_req = load_workbook(excel_path)

    tables = ['primMon', 'primOmon', 'primDay', 'primOday', 'primSIday',
              'primO6hr', 'prim6hr', 'prim6hrpt', 'prim3hr', 'prim3hrpt',
              'prim1hrpt']

    for table in tables:
        # create a blank output dictionary
        output = {}

        # choose the corresponding worksheet from the Excel data request
        req_sheet = data_req[table]

        # add the Header dictionary
        output['Header'] = generate_header(table)

        # add the variables
        output['variable_entry'] = {}
        generate_variable_entry(req_sheet, output['variable_entry'])

        # update the header with the realms from the variables
        unique_realms = {output['variable_entry'][var]['modeling_realm']: 1
                         for var in output['variable_entry']}
        output['Header']['realm'] = ' '.join(unique_realms.keys())

        # write the new JSON file for the PRIMAVERA table
        output_json_name = 'CMIP6_{}.json'.format(table)
        output_path = os.path.abspath(
            os.path.join(current_dir, '..', 'Tables', output_json_name))
        with open(output_path, 'w') as dest_file:
            json.dump(output, dest_file, indent=4)


def _get_cell(row, column_name):
    """
    Return the contents of the specified column from the row passed in.

    :param tuple row: a row (a tuple of cell objects) as returned by
        Worksheet.iter_rows()
    :param str column_name: the name of the column
    :returns: cell.value
    """
    column_names = {'long_name': 1, 'units': 2, 'comment': 3,
                    'var_name': 5, 'standard_name': 6, 'cell_methods': 7,
                    'positive': 8, 'type': 9, 'dimensions': 10,
                    'cmor_name': 11, 'modeling_realm': 12, 'frequency': 13,
                    'cell_measures': 14}

    return row[column_names[column_name]].value


if __name__ == '__main__':
    main()
